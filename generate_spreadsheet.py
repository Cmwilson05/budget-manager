#!/usr/bin/env python3
"""Generate the Budget Manager Excel spreadsheet."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

wb = Workbook()

# ============================================================
# SHEET 1: ACCOUNTS
# ============================================================
ws_accounts = wb.active
ws_accounts.title = "Accounts"

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
liability_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
summary_font = Font(bold=True, size=12)
money_format = '#,##0.00'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column widths
ws_accounts.column_dimensions['A'].width = 25
ws_accounts.column_dimensions['B'].width = 15
ws_accounts.column_dimensions['C'].width = 12
ws_accounts.column_dimensions['D'].width = 20

# Header row
headers = ["Name", "Balance", "Type", "Include in Workbench?"]
for col, h in enumerate(headers, 1):
    cell = ws_accounts.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Sample data (rows 2-8, user will replace)
sample_accounts = [
    ("Checking Account", 4500.00, "Asset", True),
    ("Savings Account", 12000.00, "Asset", True),
    ("Venmo", 350.00, "Asset", True),
    ("Credit Card - CMW (3619)", 1200.00, "Liability", True),
    ("Credit Card - JGW (9299)", 800.00, "Liability", True),
    ("Credit Card - SAMS (1261)", 450.00, "Liability", True),
]

for i, (name, balance, acct_type, include) in enumerate(sample_accounts, 2):
    ws_accounts.cell(row=i, column=1, value=name).border = thin_border
    c = ws_accounts.cell(row=i, column=2, value=balance)
    c.number_format = money_format
    c.border = thin_border
    ws_accounts.cell(row=i, column=3, value=acct_type).border = thin_border
    ws_accounts.cell(row=i, column=4, value=include).border = thin_border
    if acct_type == "Liability":
        for col in range(1, 5):
            ws_accounts.cell(row=i, column=col).fill = liability_fill

# Data validation for Type column
dv_type = DataValidation(type="list", formula1='"Asset,Liability"', allow_blank=False)
dv_type.error = "Must be Asset or Liability"
ws_accounts.add_data_validation(dv_type)
dv_type.add(f"C2:C100")

# Data validation for Include column
dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
ws_accounts.add_data_validation(dv_bool)
dv_bool.add(f"D2:D100")

# Summary section (row 9+, below data)
summary_start = len(sample_accounts) + 3  # leave a blank row

ws_accounts.cell(row=summary_start, column=1, value="Total Assets").font = summary_font
ws_accounts.cell(row=summary_start, column=2).font = summary_font
ws_accounts.cell(row=summary_start, column=2).number_format = money_format
ws_accounts.cell(row=summary_start, column=2, value=f'=SUMIFS(B2:B100,C2:C100,"Asset")')

ws_accounts.cell(row=summary_start+1, column=1, value="Total Liabilities").font = summary_font
ws_accounts.cell(row=summary_start+1, column=2).font = summary_font
ws_accounts.cell(row=summary_start+1, column=2).number_format = money_format
ws_accounts.cell(row=summary_start+1, column=2, value=f'=SUMIFS(B2:B100,C2:C100,"Liability")')

ws_accounts.cell(row=summary_start+2, column=1, value="Net Worth").font = Font(bold=True, size=14)
ws_accounts.cell(row=summary_start+2, column=2).font = Font(bold=True, size=14)
ws_accounts.cell(row=summary_start+2, column=2).number_format = money_format
ws_accounts.cell(row=summary_start+2, column=2, value=f'=B{summary_start}-B{summary_start+1}')

ws_accounts.cell(row=summary_start+4, column=1, value="Workbench Starting Balance").font = Font(bold=True, size=12, color="1F4E79")
ws_accounts.cell(row=summary_start+4, column=2).font = Font(bold=True, size=12, color="1F4E79")
ws_accounts.cell(row=summary_start+4, column=2).number_format = money_format
ws_accounts.cell(row=summary_start+4, column=2, value=f'=SUMIFS(B2:B100,C2:C100,"Asset",D2:D100,TRUE)-SUMIFS(B2:B100,C2:C100,"Liability",D2:D100,TRUE)')

# Name the workbench balance cell for cross-sheet reference
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names.add(DefinedName("WorkbenchBalance", attr_text=f"Accounts!$B${summary_start+4}"))

# ============================================================
# SHEET 2: MAIN WORKBENCH
# ============================================================
ws_main = wb.create_sheet("Main Workbench")

# Column widths
ws_main.column_dimensions['A'].width = 12
ws_main.column_dimensions['B'].width = 14
ws_main.column_dimensions['C'].width = 30
ws_main.column_dimensions['D'].width = 15

# Summary header section
dark_fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
dark_font = Font(bold=True, color="FFFFFF", size=10)
dark_value_font = Font(bold=True, color="FFFFFF", size=14)
green_font = Font(bold=True, color="00B894", size=14)
red_font = Font(bold=True, color="D63031", size=14)
blue_font = Font(bold=True, color="74B9FF", size=14)

# Row 1: Title
ws_main.merge_cells('A1:D1')
title_cell = ws_main.cell(row=1, column=1, value="Main Cash Flow")
title_cell.font = Font(bold=True, size=16, color="FFFFFF")
title_cell.fill = dark_fill
title_cell.alignment = Alignment(horizontal='center')
for col in range(1, 5):
    ws_main.cell(row=1, column=col).fill = dark_fill

# Row 2-3: Summary metrics
labels = ["Current", "Income", "Expenses", "Projected"]
for col, label in enumerate(labels, 1):
    c = ws_main.cell(row=2, column=col, value=label)
    c.font = dark_font
    c.fill = dark_fill
    c.alignment = Alignment(horizontal='center')

# Formulas (data starts row 6)
DATA_START = 6
DATA_END = 200

# Current balance - reference Accounts sheet
ws_main.cell(row=3, column=1, value=f"=Accounts!B{summary_start+4}")
ws_main.cell(row=3, column=1).font = dark_value_font
ws_main.cell(row=3, column=1).fill = dark_fill
ws_main.cell(row=3, column=1).number_format = '$#,##0.00'
ws_main.cell(row=3, column=1).alignment = Alignment(horizontal='center')

# Income (sum positive amounts where InCalc=TRUE)
ws_main.cell(row=3, column=2, value=f'=SUMPRODUCT((A{DATA_START}:A{DATA_END}=TRUE)*(D{DATA_START}:D{DATA_END}>0)*D{DATA_START}:D{DATA_END})')
ws_main.cell(row=3, column=2).font = green_font
ws_main.cell(row=3, column=2).fill = dark_fill
ws_main.cell(row=3, column=2).number_format = '$#,##0.00'
ws_main.cell(row=3, column=2).alignment = Alignment(horizontal='center')

# Expenses (sum negative amounts where InCalc=TRUE)
ws_main.cell(row=3, column=3, value=f'=SUMPRODUCT((A{DATA_START}:A{DATA_END}=TRUE)*(D{DATA_START}:D{DATA_END}<0)*D{DATA_START}:D{DATA_END})')
ws_main.cell(row=3, column=3).font = red_font
ws_main.cell(row=3, column=3).fill = dark_fill
ws_main.cell(row=3, column=3).number_format = '$#,##0.00'
ws_main.cell(row=3, column=3).alignment = Alignment(horizontal='center')

# Projected (current + income + expenses)
ws_main.cell(row=3, column=4, value='=A3+B3+C3')
ws_main.cell(row=3, column=4).font = blue_font
ws_main.cell(row=3, column=4).fill = dark_fill
ws_main.cell(row=3, column=4).number_format = '$#,##0.00'
ws_main.cell(row=3, column=4).alignment = Alignment(horizontal='center')

# Row 4: blank spacer
# Row 5: Column headers for transaction data
headers = ["In Calc?", "Due Date", "Description", "Amount"]
for col, h in enumerate(headers, 1):
    cell = ws_main.cell(row=5, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Sample transactions
sample_transactions = [
    (True, date(2026, 6, 1), "Rent / Mortgage", -1800.00),
    (True, date(2026, 6, 5), "Paycheck", 2800.00),
    (True, date(2026, 6, 12), "Electric Bill", -145.00),
    (True, date(2026, 6, 15), "Paycheck", 2800.00),
    (False, date(2026, 6, 20), "New Furniture (planning)", -600.00),
    (True, date(2026, 6, 25), "Internet", -79.99),
]

for i, (in_calc, due, desc, amount) in enumerate(sample_transactions, DATA_START):
    ws_main.cell(row=i, column=1, value=in_calc).border = thin_border
    ws_main.cell(row=i, column=1).alignment = Alignment(horizontal='center')
    c = ws_main.cell(row=i, column=2, value=due)
    c.number_format = 'M/D/YY'
    c.border = thin_border
    ws_main.cell(row=i, column=3, value=desc).border = thin_border
    amt_cell = ws_main.cell(row=i, column=4, value=amount)
    amt_cell.number_format = '$#,##0.00'
    amt_cell.border = thin_border
    amt_cell.font = Font(color="00B894" if amount >= 0 else "D63031")

# Data validation for In Calc column
dv_calc = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
ws_main.add_data_validation(dv_calc)
dv_calc.add(f"A{DATA_START}:A{DATA_END}")

# Conditional formatting: gray out rows where In Calc = FALSE
gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ws_main.conditional_formatting.add(
    f"A{DATA_START}:D{DATA_END}",
    FormulaRule(formula=[f'$A{DATA_START}=FALSE'], fill=gray_fill)
)

# ============================================================
# SHEET 3: CC WORKBENCHES
# ============================================================
cc_accounts = [
    ("CC - CMW (3619)", "cc_1", "Credit Card - CMW (3619)"),
    ("CC - JGW (9299)", "cc_2", "Credit Card - JGW (9299)"),
    ("CC - SAMS (1261)", "cc_3", "Credit Card - SAMS (1261)"),
]

for sheet_name, tag, account_name in cc_accounts:
    ws_cc = wb.create_sheet(sheet_name)

    ws_cc.column_dimensions['A'].width = 12
    ws_cc.column_dimensions['B'].width = 14
    ws_cc.column_dimensions['C'].width = 30
    ws_cc.column_dimensions['D'].width = 15

    # Title
    ws_cc.merge_cells('A1:D1')
    title_cell = ws_cc.cell(row=1, column=1, value=account_name)
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = dark_fill
    title_cell.alignment = Alignment(horizontal='center')
    for col in range(1, 5):
        ws_cc.cell(row=1, column=col).fill = dark_fill

    # Row 2: labels
    labels = ["Balance Owed", "Charges", "Payments", "Projected"]
    for col, label in enumerate(labels, 1):
        c = ws_cc.cell(row=2, column=col, value=label)
        c.font = dark_font
        c.fill = dark_fill
        c.alignment = Alignment(horizontal='center')

    # Balance owed - VLOOKUP from Accounts
    ws_cc.cell(row=3, column=1, value=f'=IFERROR(-SUMIFS(Accounts!B2:B100,Accounts!A2:A100,"*{account_name.split(" - ")[1].split(" ")[0]}*",Accounts!C2:C100,"Liability"),0)')
    ws_cc.cell(row=3, column=1).font = red_font
    ws_cc.cell(row=3, column=1).fill = dark_fill
    ws_cc.cell(row=3, column=1).number_format = '$#,##0.00'
    ws_cc.cell(row=3, column=1).alignment = Alignment(horizontal='center')

    # Charges (negative amounts where in_calc=TRUE)
    ws_cc.cell(row=3, column=2, value=f'=SUMPRODUCT((A{DATA_START}:A{DATA_END}=TRUE)*(D{DATA_START}:D{DATA_END}<0)*D{DATA_START}:D{DATA_END})')
    ws_cc.cell(row=3, column=2).font = red_font
    ws_cc.cell(row=3, column=2).fill = dark_fill
    ws_cc.cell(row=3, column=2).number_format = '$#,##0.00'
    ws_cc.cell(row=3, column=2).alignment = Alignment(horizontal='center')

    # Payments (positive amounts where in_calc=TRUE)
    ws_cc.cell(row=3, column=3, value=f'=SUMPRODUCT((A{DATA_START}:A{DATA_END}=TRUE)*(D{DATA_START}:D{DATA_END}>0)*D{DATA_START}:D{DATA_END})')
    ws_cc.cell(row=3, column=3).font = green_font
    ws_cc.cell(row=3, column=3).fill = dark_fill
    ws_cc.cell(row=3, column=3).number_format = '$#,##0.00'
    ws_cc.cell(row=3, column=3).alignment = Alignment(horizontal='center')

    # Projected
    ws_cc.cell(row=3, column=4, value='=A3+B3+C3')
    ws_cc.cell(row=3, column=4).font = blue_font
    ws_cc.cell(row=3, column=4).fill = dark_fill
    ws_cc.cell(row=3, column=4).number_format = '$#,##0.00'
    ws_cc.cell(row=3, column=4).alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["In Calc?", "Due Date", "Description", "Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws_cc.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data validation
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws_cc.add_data_validation(dv)
    dv.add(f"A{DATA_START}:A{DATA_END}")

    # Gray out inactive rows
    ws_cc.conditional_formatting.add(
        f"A{DATA_START}:D{DATA_END}",
        FormulaRule(formula=[f'$A{DATA_START}=FALSE'], fill=gray_fill)
    )


# ============================================================
# SHEET 4: BILL SCHEDULE
# ============================================================
ws_bills = wb.create_sheet("Bill Schedule")

ws_bills.column_dimensions['A'].width = 25
ws_bills.column_dimensions['B'].width = 12
ws_bills.column_dimensions['C'].width = 14
ws_bills.column_dimensions['D'].width = 14
ws_bills.column_dimensions['E'].width = 14

# Headers
bill_headers = ["Name", "Amount", "Frequency", "Next Due", "Last Paid"]
for col, h in enumerate(bill_headers, 1):
    cell = ws_bills.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Sample data
sample_bills = [
    ("Mortgage", 1800.00, "Monthly", date(2026, 6, 1), date(2026, 5, 1)),
    ("Electric", 145.00, "Monthly", date(2026, 6, 12), date(2026, 5, 12)),
    ("Internet", 79.99, "Monthly", date(2026, 6, 25), date(2026, 5, 25)),
    ("Netflix", 15.99, "Monthly", date(2026, 6, 8), date(2026, 5, 8)),
    ("Spotify", 11.99, "Monthly", date(2026, 6, 15), date(2026, 5, 15)),
    ("Car Insurance", 480.00, "Quarterly", date(2026, 7, 15), date(2026, 4, 15)),
    ("Amazon Prime", 139.00, "Annually", date(2026, 11, 20), date(2025, 11, 20)),
    ("Paycheck", 2800.00, "Bi-Weekly", date(2026, 6, 5), date(2026, 5, 22)),
]

for i, (name, amount, freq, due, paid) in enumerate(sample_bills, 2):
    ws_bills.cell(row=i, column=1, value=name).border = thin_border
    c = ws_bills.cell(row=i, column=2, value=amount)
    c.number_format = money_format
    c.border = thin_border
    ws_bills.cell(row=i, column=3, value=freq).border = thin_border
    c = ws_bills.cell(row=i, column=4, value=due)
    c.number_format = 'M/D/YY'
    c.border = thin_border
    c = ws_bills.cell(row=i, column=5, value=paid)
    c.number_format = 'M/D/YY'
    c.border = thin_border

# Frequency validation
dv_freq = DataValidation(type="list", formula1='"Bi-Weekly,Monthly,Quarterly,Annually"', allow_blank=False)
ws_bills.add_data_validation(dv_freq)
dv_freq.add("C2:C200")

# Conditional formatting: past due (red) and due soon (yellow)
red_fill_cf = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
yellow_fill_cf = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

# Past due: Next Due <= TODAY()
ws_bills.conditional_formatting.add(
    "A2:E200",
    FormulaRule(formula=['AND($D2<>"", $D2<=TODAY())'], fill=red_fill_cf)
)

# Due within 7 days: Next Due > TODAY() and Next Due <= TODAY()+7
ws_bills.conditional_formatting.add(
    "A2:E200",
    FormulaRule(formula=['AND($D2<>"", $D2>TODAY(), $D2<=TODAY()+7)'], fill=yellow_fill_cf)
)

# Summary row
bill_summary_row = len(sample_bills) + 3
ws_bills.cell(row=bill_summary_row, column=1, value="Total Monthly Fixed").font = summary_font
ws_bills.cell(row=bill_summary_row, column=2, value='=SUMIFS(B2:B200,C2:C200,"<>Annually")').font = summary_font
ws_bills.cell(row=bill_summary_row, column=2).number_format = money_format

ws_bills.cell(row=bill_summary_row+1, column=1, value="Safe to Spend").font = Font(bold=True, size=12, color="1F4E79")
ws_bills.cell(row=bill_summary_row+1, column=2, value=f'=SUMIFS(Accounts!B2:B100,Accounts!C2:C100,"Asset")-B{bill_summary_row}')
ws_bills.cell(row=bill_summary_row+1, column=2).font = Font(bold=True, size=12, color="1F4E79")
ws_bills.cell(row=bill_summary_row+1, column=2).number_format = money_format


# ============================================================
# SAVE
# ============================================================
output_path = "/home/user/budget-manager/budget-manager.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
